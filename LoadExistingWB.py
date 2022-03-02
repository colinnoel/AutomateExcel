from openpyxl import Workbook, load_workbook

wb = load_workbook('Sales.xlsx')
ws = wb.active # gives us the active worksheet from the loaded workbook
# ws = wb['Sheet1'] <- access a different sheet in the workbook

print(ws['A2'].value)

ws['A9'].value = "Test"  
wb.create_sheet("NewSheet")

wb.save('Sales.xlsx')

print(wb.sheetnames) #<- View the different sheets in a workbook

