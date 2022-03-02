from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('SalesWBCreatedInPython.xlsx')
ws = wb.active

ws.merge_cells("A10:D10")


wb.save('SalesWBCreatedInPython.xlsx')