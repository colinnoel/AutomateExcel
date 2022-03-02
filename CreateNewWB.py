from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = Workbook() <- Creating a new workbook from the .py file
# ws.title = "SalesData"

wb = load_workbook('SalesWBCreatedInPython.xlsx')
ws = wb.active

ws.append(['Sales Rep', 'Q1', 'Q2', 'Q3', 'Q4'])

for row in range(2,11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)


wb.save('SalesWBCreatedInPython.xlsx')
